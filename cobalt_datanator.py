
from openpyxl import  Workbook
import xml.etree.cElementTree as ET
import argparse
import os
from operator import itemgetter
import datetime
import sqlite3
import re
import dateutil.parser

import pickle
import pytz

def dict_to_ordered_array_for_excel(dict, sortkey):
    for entry in dict:
        if isinstance(dict[entry][sortkey],str):
            try:
                dict[entry][sortkey] = datetime.datetime.fromtimestamp(int(dict[entry][sortkey]) / 1000   )
            except:
                print("we already a datetime up in this dog")
    sorted_array = sorted(list(dict.values()), key=itemgetter(sortkey))
    return sorted_array






class CobaltData:

    ##used to filter out users or hosts
    filter = {
        "user": ["sandbox/testuser"]
    }

    def __init__(self, cobalt_dir=False, empire_dir=False, covenant_dir=False):
        self.cobalt_dir = cobalt_dir
        if cobalt_dir:
            self.sessionsxml,self.activityxml,self.c2xml,self.credentialsxml = self.load_cobalt_data_from_dir()
            self.cobalt_sessions, self.cobalt_c2, self.cobalt_activities, self.cobalt_credentials = self.parse_cobalt_xml()

        self.empire_dir = empire_dir
        if empire_dir:
            self.empire_sessions, self.empire_activities = self.parse_empire_data()

        self.covenant_dir = covenant_dir
        if covenant_dir:
            self.covenant_sessions, self.covenant_activities = self.parse_covenant_data()

        self.sessions = {}
        self.activity = []
        self.credentials = []
        self.combinate_da_c2()






    def combinate_da_c2(self):
        if self.cobalt_dir:
            for key, session in self.cobalt_sessions.items():
                session["agent_type"] = "cobalt_strike"
                self.sessions[key] = session
            for activity in self.cobalt_activities:
                activity["agent_type"] = "cobalt_strike"
                self.activity.append(activity)
            for credential in self.cobalt_credentials:
                credential["agent_type"] = "cobalt_strike"
                self.credentials.append(credential)

        if self.empire_dir:
            for key, session in self.empire_sessions.items():
                session["agent_type"] = "empire"
                self.sessions[key] = session
            for activity in self.empire_activities:
                activity["agent_type"] = "empire"
                self.activity.append(activity)

        if self.covenant_dir:
            for key, session in self.covenant_sessions.items():
                session["agent_type"] = "covenant"
                self.sessions[key] = session
            for activity in self.covenant_activities:
                activity["agent_type"] = "covenant"
                self.activity.append(activity)


    def xml_entries_to_dict(self,xml,primarykey,exclude_keys=[]):
        return_dict = {}
        filter_dict = self.filter
        for entry in xml:
            entry_dict = {}
            filtered = False
            for attrib in entry:
                if attrib.tag == primarykey:
                    primarykey_value = attrib.text
                if attrib.tag in filter_dict:
                    for filter in filter_dict[attrib.tag]:
                        if filter in attrib.text:
                            filtered = True
                if attrib.tag in exclude_keys:
                    continue
                entry_dict[attrib.tag] = attrib.text
            if not filtered:
                return_dict[primarykey_value] = entry_dict
        return return_dict

    def xml_entries_to_list(self,xml,exclude_keys=[]):
        return_array = []
        filter_dict = self.filter
        for entry in xml:
            entry_dict = {}
            filtered = False
            for attrib in entry:
                if attrib.tag in filter_dict:
                    for filter in filter_dict[attrib.tag]:
                        if filter in attrib.text:
                            filtered = True
                if attrib.tag in exclude_keys:
                    continue
                entry_dict[attrib.tag] = attrib.text
            if not filtered:
                return_array.append(entry_dict)
        return return_array

    def parse_cobalt_xml(self):

        sessions = self.xml_entries_to_dict(self.sessionsxml,"id",["is64"])
        c2 = self.xml_entries_to_dict(self.c2xml,"bid",["bid"])
        activities = self.xml_entries_to_list(self.activityxml,["tactic"])
        credentials = self.xml_entries_to_list(self.credentialsxml,["note"])


        for session in sessions:
            if session in c2:
                sessions[session]["opened"] = datetime.datetime.fromtimestamp(int(sessions[session]["opened"] ) / 1000)

                for c2key in c2[session]:
                    sessions[session][c2key] = c2[session][c2key]



        session_fields_to_include_in_activity = [
            "user", "internal", "computer","pid","domains","port"
        ]
        activities_with_sessions = []
        for activity in activities:
            activity["when"] = datetime.datetime.fromtimestamp(int(activity["when"]) / 1000 )
            if activity["bid"] in sessions:
                for sessionkey in session_fields_to_include_in_activity:
                    activity[sessionkey] = sessions[activity["bid"]][sessionkey]
                activities_with_sessions.append(activity)


        for credential in credentials:
            for session in sessions.values():
                if credential["host"]  == session["internal"]:
                    credential["host"] = session["computer"]


        return sessions, c2, activities_with_sessions, credentials






    def load_xml_root(self, file):
        with open(file,'r') as xmlfile:
            xml =xmlfile.read()
            xml = xml.replace('\0','')
            xml = xml.replace('&','&amp;')
            xml = xml.replace("\u001f", " ")
            tree = ET.fromstring(xml)
            return tree

        ## unlike cobalt data load, this function returns python dicts/arrays
    def parse_empire_data(self):
        empire_dir = self.empire_dir

        if not os.path.isdir(empire_dir):
            raise FileNotFoundError("not a directory")
            print("not a directory")

        data_dir = os.path.join(empire_dir, 'data')
        if not os.path.isdir(data_dir):
            raise FileNotFoundError("data dir not found")
            print("data dir not found")

        downloads_dir = os.path.join(empire_dir, 'downloads')
        if not os.path.isdir(downloads_dir):
            raise FileNotFoundError("downloads dir not found")
            print("downloads dir not found")

        ##load in data from Empire database
        empire_db_file = os.path.join(data_dir,'empire.db')
        if not os.path.exists(empire_db_file):
            raise FileNotFoundError("empire db not found")
            print("empire db  not found")

        conn = sqlite3.connect(empire_db_file)
        conn.row_factory = sqlite3.Row

        c = conn.cursor()

        #collect sessions
        c.execute("select session_id as id, checkin_time as opened,  external_ip as external, internal_ip as internal, agents.username as user, high_integrity, hostname as computer, process_id as pid, parent as ppid, options as options from agents inner join listeners on agents.listener = listeners.name")
        sessions = {}
        for row in c:
            id  = row["id"]
            session = dict(zip(row.keys(), row))
            if session["high_integrity"] ==1:
                session["user"] = session["user"] + "*"
            session.pop("high_integrity")

            options_pickle = session.pop("options")
            print("Warning you are being pickled, this might get you pwned YOLO")
            options = pickle.loads(str.encode(options_pickle))
            session["port"] = options["Port"]["Value"]
            session["domains"] = options["Host"]["Value"]

            session["opened"] =   datetime.datetime.strptime( session["opened"],'%Y-%m-%d %H:%M:%S')
            sessions[id] = session






        activities = []
        ## time to load in the files


        for session in sessions:
            agent_file = os.path.join(downloads_dir,session,'agent.log')
            if os.path.exists(agent_file):
                with open(agent_file,'r', encoding="ISO-8859-1") as f:
                    line = f.readline()
                    while line:
                        r = re.match("(\d\d\d\d-\d\d-\d\d\s\d\d:\d\d:\d\d)",line)
                        if r is not None:
                            activity  = {}

                            date_string = r.group(1)
                            activity["when"] = datetime.datetime.strptime(date_string,'%Y-%m-%d %H:%M:%S')
                            line = f.readline()
                            while line and not (line.startswith(("Tasked", "[*] Tasked", "[+] Agent"))) \
                                    and not re.match("(\d\d\d\d-\d\d-\d\d\s\d\d:\d\d:\d\d)",line):
                                line = f.readline()
                            if re.match("(\d\d\d\d-\d\d-\d\d\s\d\d:\d\d:\d\d)",line):
                                continue
                            if not line:
                                continue
                            activity["data"] = line


                            if line.startswith("[+] Agent"):
                                activity["type"] = "initial"
                            else:
                                activity["type"]= "task"

                            activity["bid"] = session
                            session_dict = sessions[session]
                            session_fields_to_include_in_activity = [
                                "user", "internal", "computer", "pid", "domains", "port"
                            ]
                            activities_with_sessions = []
                            for sessionkey in session_fields_to_include_in_activity:
                                activity[sessionkey] = session_dict[sessionkey]
                            activities.append(activity)
                        line = f.readline()

        return sessions, activities

    def parse_covenant_data(self):
        covenant_dir = self.covenant_dir

        if not os.path.isdir(covenant_dir):
            raise FileNotFoundError("not a directory")
            print("not a directory")

        data_dir = os.path.join(covenant_dir, 'data')
        if not os.path.isdir(data_dir):
            raise FileNotFoundError("data dir not found")
            print("data dir not found")


        ##load in data from covenant database
        db_file = os.path.join(covenant_dir,'data', 'covenant.db')
        if not os.path.exists(db_file):
            raise FileNotFoundError("covenant db not found")
            print("covenant db  not found")

        conn = sqlite3.connect(db_file)
        conn.row_factory = sqlite3.Row

        c = conn.cursor()

        # collect grunts and commands
        c.execute(
            "select  Grunts.id, ActivationTime as opened,  ConnectAddresses as external, IPAddress as internal, UserName as user, Integrity as high_integrity, HostName as computer, Process as pid  from Grunts inner join Listeners on Grunts.listenerid = Listeners.id;")
        sessions = {}
        for row in c:
            id =  "cov-" + str(row["id"])
            session = dict(zip(row.keys(), row))
            if session["high_integrity"] == 1:
                session["user"] = session["user"] + "*"
            session.pop("high_integrity")


            #harcoded for kevin
            session["port"] = "443"
            session["domains"] = "onedrive.com, storage.live.com"

            session["opened"] = dateutil.parser.parse(session["opened"])
            sessions[id] = session

        activities = []
        ## time to load in the files

        session_fields_to_include_in_activity = [
            "user", "internal", "computer", "pid", "domains", "port"
        ]
        c.execute("select  GruntId as bid, CommandTime as \"when\",  Command as data from GruntCommands;")
        for row in c:
            bid = "cov-" + str(row["bid"])
            session_dict = sessions[bid]
            activity =dict(zip(row.keys(), row))
            activity["type"] = "covenant command"
            activity["when"] =  dateutil.parser.parse(activity["when"])
            for sessionkey in session_fields_to_include_in_activity:
                activity[sessionkey] = session_dict[sessionkey]
            activities.append(activity)


        return sessions, activities










    def load_cobalt_data_from_dir(self):



        cobalt_dir = self.cobalt_dir

        COBALT_FILES = {
            "sessionsxmlfile": "sessions.xml",
            "activityxmlfile": "activity.xml",
            "c2xmlfile": "c2info.xml",
            "credentialsxmlfile": "credentials.xml"
        }



        if not os.path.isdir(cobalt_dir):
            raise FileNotFoundError("not a directory")
            print("not a directory")



        cxml = {}
        for cfile in COBALT_FILES:
            xmlpath = os.path.join(cobalt_dir,COBALT_FILES[cfile])
            if not (os.path.exists(xmlpath)):
                raise FileNotFoundError("missing cobalt xml export files. Ensure all exported cobalt XML files are in the specified dir")
            cxml[cfile] = self.load_xml_root(xmlpath)


        return cxml["sessionsxmlfile"], cxml["activityxmlfile"], cxml["c2xmlfile"], cxml["credentialsxmlfile"]



    def make_sessions_report(self):

        sessions_array = dict_to_ordered_array_for_excel(self.sessions, "opened")

        wb = Workbook()
        ws = wb.active
        ws.title= "Sessions"

        header = sessions_array[0].keys()

        ws.append(list(header))

        for session in sessions_array:
            row = []
            for head in header:
                if head not in session:
                    session[head] = ""
                row.append(session[head])
            ws.append(row)
        ws.auto_filter.ref = 'A:{}'.format(chr(ord('A') + len(header) -1))




        ws =wb.create_sheet("Activity")

        activity_array = sorted(self.activity, key=itemgetter("when"))

        header = activity_array[0].keys()
        ws.append(list(header))

        row = 0
        for activity in activity_array:
            col = 0
            row = []
            for head in header:
                row.append(activity[head])
            ws.append(row)



        ws.auto_filter.ref = 'A:{}'.format(chr(ord('A') + len(header) -1))


#credz
        try:
            ws = wb.create_sheet("Credentials")

            credentials_array = self.credentials
            header = credentials_array[0].keys()
            ws.append(list(header))

            for activity in credentials_array:
                ws.append(list(activity.values()))

            ws.auto_filter.ref = 'A:{}'.format(chr(ord('A') + len(header) - 1))
        except:
            pass
        return wb




if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", help="Directory containing cobalt data", required=False)
    parser.add_argument("-e", help="Directory containing empire data", required=False)
    parser.add_argument("-c", help="Directory containing cov data", required=False)
    parser.add_argument("-o", help="Output file", required=True)

    parser.add_argument("--filter", help="filter", required=False)



    args = parser.parse_args()

    cobalt_dir = args.d
    empire_dir = args.e
    covenant_dir = args.c
    cobalt_data = CobaltData(cobalt_dir, empire_dir, covenant_dir)





    sessions_workbook = cobalt_data.make_sessions_report()

    sessions_workbook.save(args.o)





